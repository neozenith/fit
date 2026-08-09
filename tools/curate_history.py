#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl==3.1.5", "polars==1.34.0"]
# ///
"""Curate the historical tracker workbook into typed Parquet.

The source is a Google Forms responses sheet grown into a 32-sheet workbook.
Only FOUR of those sheets are facts; the other twenty-eight are derived —
pivots, dashboards, calendars, forecasts. This script reads the facts and
ignores every derivation, because the derivations are what the app is going to
recompute. Curating a pivot table would bake one spreadsheet's arithmetic into
the data layer and make the app's own numbers unverifiable against anything.

    WorkoutDataClean   -> strength_sets.parquet     one row per set-group
    BodyWeightClean    -> body_metrics.parquet      one row per weigh-in
    StravaActivities   -> cardio_activities.parquet one row per activity
    (derived)          -> exercises.parquet         the exercise catalogue

Output goes to reference/history/, which is gitignored along with the rest of
reference/: this is personal body-composition data and it stays on the machine
that produced it. Publishing to an environment is a separate, explicit step
(tools/publish-history.ts).

Not degradable. Every dependency is imported at the top and every sheet is
required — a workbook missing one of them is a workbook this script cannot
honestly curate, so it raises rather than emitting a partial dataset.
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook

log = logging.getLogger("curate_history")

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = ROOT / "reference" / "OLD Workout Performance Tracker (Responses).xlsx"
DEFAULT_OUT = ROOT / "reference" / "history"

# Exercises whose "reps" column is SECONDS HELD, not repetitions. Multiplying
# those by a load would produce a volume number in units nothing else shares,
# and it would dominate every chart it appeared in.
ISOMETRIC = re.compile(r"\(time\)$", re.IGNORECASE)

# Exercises loaded by the body rather than by a bar. Their recorded "weight" is
# the athlete's own mass, which is why it tracks body weight over time instead
# of stepping in plate increments.
BODYWEIGHT_LOADED = {"Pull up", "Push ups", "Push ups (feet elevated)"}

# The form used ONE question for everything, so a weigh-in is recorded as an
# "exercise" of 1 set x 1 rep at body weight — and those rows sit in the
# strength sheet alongside real lifts. Left in, they are half the strength log:
# 498 phantom sessions, 44 tonnes of phantom volume, and a 42nd "exercise"
# called Body Weight sitting third in the volume ranking. They are already
# curated properly into body_metrics, so here they are excluded.
WEIGH_IN_EXERCISE = "Body Weight"

# One arm or one leg at a time — the recorded volume is per side, so a like for
# like comparison against a bilateral lift has to know.
UNILATERAL = re.compile(r"\b(SA|Single Arm|Single Leg|each side)\b", re.IGNORECASE)

KNOWN_EQUIPMENT = (
    "Barbell",
    "Dumbbell",
    "Kettlebell",
    "Cable",
    "Landmine",
    "Plate",
    "Machine",
)


def _sheet_rows(workbook: Any, name: str) -> tuple[tuple[Any, ...], list[tuple[Any, ...]]]:
    """Header plus non-blank data rows.

    `read_only` worksheets report no dimensions until forced, and the sheets
    here are padded to 1000+ rows with formula cells that evaluate to blank —
    so "non-blank" has to mean "some cell holds something", not "the row exists".
    """
    if name not in workbook.sheetnames:
        raise SystemExit(f"error: the workbook has no sheet named {name!r}")
    sheet = workbook[name]
    sheet.calculate_dimension(force=True)
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    data = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    return header, data


def _num(value: Any) -> float | None:
    """Coerce a spreadsheet cell to a float, or None.

    Everything arrives as a string because the sheet was exported with
    `data_only`, and a share of the cells are spreadsheet error literals
    (`#VALUE!`, `#N/A`) left behind by formulas over empty ranges. Those are
    absence, not zero — a `#VALUE!` scored as 0 would drag every average down.
    """
    if value is None:
        return None
    text = str(value).strip()
    if text == "" or text.startswith("#"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _timestamp(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    # Google Forms writes microseconds; some rows were hand-edited and lost them.
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _epoch_ms(value: Any) -> datetime | None:
    """Strava start dates arrive as epoch milliseconds, as floats."""
    ms = _num(value)
    if ms is None:
        return None
    return datetime.fromtimestamp(ms / 1000.0)


def curate_strength(workbook: Any) -> tuple[pl.DataFrame, int, int]:
    _, rows = _sheet_rows(workbook, "WorkoutDataClean")
    records = []
    skipped = 0
    weigh_ins = 0
    for date, exercise, sets, reps, weight in ((r + (None,) * 5)[:5] for r in rows):
        when = _timestamp(date)
        name = str(exercise).strip() if exercise else ""
        n_sets, n_reps, kg = _num(sets), _num(reps), _num(weight)
        if name == WEIGH_IN_EXERCISE:
            weigh_ins += 1
            continue
        if when is None or not name or n_sets is None or n_reps is None or kg is None:
            # The sheet is padded with formula rows that evaluate to
            # (blank, blank, 1, 1, 0) — 100 of them. They are filler, not
            # entries, but the count is REPORTED rather than quietly dropped:
            # a silent skip is indistinguishable from a parser that broke.
            skipped += 1
            continue
        isometric = bool(ISOMETRIC.search(name))
        records.append(
            {
                "logged_at": when,
                "date": when.date(),
                "exercise": name,
                "sets": int(n_sets),
                "reps": int(n_reps),
                "weight_kg": kg,
                # Volume is undefined for a hold: `reps` there is seconds, and
                # sets x seconds x kg is not a load. Left null so a SUM skips it
                # rather than a 0 that looks like a session with no work in it.
                "volume_kg": None if isometric else n_sets * n_reps * kg,
                "is_isometric": isometric,
            }
        )
    return pl.DataFrame(records).sort("logged_at"), skipped, weigh_ins


def curate_body(workbook: Any) -> tuple[pl.DataFrame, float]:
    """Weigh-ins, plus the height implied by the workbook's own BMI column.

    Height is never recorded anywhere in the workbook, only used. Recovering it
    from `weight / BMI` and asserting it is stable across every row is what lets
    the app compute BMI itself instead of importing a column it cannot check.
    """
    _, rows = _sheet_rows(workbook, "BodyWeightClean")
    records = []
    for date, _exercise, _sets, _reps, weight in ((r + (None,) * 5)[:5] for r in rows):
        when = _timestamp(date)
        kg = _num(weight)
        if when is None or kg is None:
            continue
        records.append({"measured_at": when, "date": when.date(), "weight_kg": kg})

    _, bmi_rows = _sheet_rows(workbook, "BodyWeightDaily")
    heights = []
    for _date, weight, _week, bmi, _month in ((r + (None,) * 5)[:5] for r in bmi_rows):
        kg, index = _num(weight), _num(bmi)
        if kg and index:
            heights.append(math.sqrt(kg / index))
    if not heights:
        raise SystemExit("error: BodyWeightDaily carries no usable weight/BMI pair")

    height_m = round(sum(heights) / len(heights), 4)
    spread = max(heights) - min(heights)
    if spread > 0.01:
        raise SystemExit(
            f"error: implied height varies by {spread:.3f}m across the workbook — "
            "the BMI column is not a pure function of weight, so it cannot be inverted"
        )

    frame = pl.DataFrame(records).sort("measured_at")
    return frame.with_columns(
        (pl.col("weight_kg") / (height_m**2)).round(2).alias("bmi")
    ), height_m


def curate_cardio(workbook: Any) -> pl.DataFrame:
    _, rows = _sheet_rows(workbook, "StravaActivities")
    records = []
    for row in rows:
        cells = (row + (None,) * 15)[:15]
        started = _epoch_ms(cells[6])
        activity_id = _num(cells[14])
        if started is None or activity_id is None:
            continue
        moving_s = _num(cells[2]) or 0.0
        distance_m = _num(cells[1]) or 0.0
        records.append(
            {
                "activity_id": int(activity_id),
                "started_at": started,
                "date": started.date(),
                "name": str(cells[0] or "").strip(),
                "activity_type": str(cells[5] or "").strip(),
                "distance_m": distance_m,
                "moving_s": int(moving_s),
                "elapsed_s": int(_num(cells[3]) or 0),
                "elevation_m": _num(cells[4]) or 0.0,
                "average_speed_ms": _num(cells[8]),
                "max_speed_ms": _num(cells[9]),
                "kilojoules": _num(cells[10]),
                "average_watts": _num(cells[11]),
                "max_watts": _num(cells[12]),
                "weighted_average_watts": _num(cells[13]),
            }
        )
    frame = pl.DataFrame(records)
    # One row per Strava id. The sheet accumulated re-imports over five years and
    # holds duplicates; a duplicated ride would double a week's distance.
    return frame.unique(subset=["activity_id"], keep="first").sort("started_at")


def classify(name: str) -> dict[str, Any]:
    equipment = next((e for e in KNOWN_EQUIPMENT if name.startswith(e)), None)
    if equipment is None:
        equipment = "Bodyweight" if name in BODYWEIGHT_LOADED or ISOMETRIC.search(name) else "Other"
    return {
        "equipment": equipment,
        "is_isometric": bool(ISOMETRIC.search(name)),
        "is_bodyweight_loaded": name in BODYWEIGHT_LOADED,
        "is_unilateral": bool(UNILATERAL.search(name)),
    }


def curate_exercises(strength: pl.DataFrame) -> pl.DataFrame:
    """The exercise catalogue, derived from what was actually performed.

    Derived rather than transcribed from the workbook's `Type` column, which is
    a formula that evaluates to `#VALUE!` on a third of its rows. A catalogue
    built from the log cannot contain an exercise that was never done, and
    cannot omit one that was.
    """
    aggregated = (
        strength.group_by("exercise")
        .agg(
            pl.col("date").min().alias("first_seen"),
            pl.col("date").max().alias("last_seen"),
            pl.len().alias("entries"),
            pl.col("sets").sum().alias("total_sets"),
            pl.col("volume_kg").sum().alias("total_volume_kg"),
            pl.col("weight_kg").max().alias("heaviest_kg"),
        )
        .sort("exercise")
    )
    traits = pl.DataFrame(
        [{"exercise": name, **classify(name)} for name in aggregated["exercise"]]
    )
    return aggregated.join(traits, on="exercise", how="left")


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    if not args.source.exists():
        raise SystemExit(f"error: {args.source} not found")

    workbook = load_workbook(args.source, read_only=True, data_only=True)

    strength, skipped, weigh_ins = curate_strength(workbook)
    body, height_m = curate_body(workbook)
    cardio = curate_cardio(workbook)
    exercises = curate_exercises(strength)

    args.out.mkdir(parents=True, exist_ok=True)
    tables = {
        "strength_sets": strength,
        "body_metrics": body,
        "cardio_activities": cardio,
        "exercises": exercises,
    }
    for name, frame in tables.items():
        if frame.is_empty():
            raise SystemExit(f"error: {name} curated to zero rows — refusing to write an empty asset")
        # A DIRECTORY per table, mirroring exactly what publish-history.ts puts
        # in S3. Identical layouts mean the same `history/{table}/**/*.parquet`
        # glob works locally and deployed, so `make dev` exercises the real
        # queries — and it leaves room to split the import by year later without
        # touching a single query.
        table_dir = args.out / name
        table_dir.mkdir(parents=True, exist_ok=True)
        frame.write_parquet(table_dir / f"{name}.parquet", compression="zstd")

    manifest = {
        "source": args.source.name,
        "curated_at": datetime.now().isoformat(timespec="seconds"),
        "height_m": height_m,
        "skipped_filler_rows": skipped,
        "weigh_in_rows_routed_to_body_metrics": weigh_ins,
        "tables": {
            name: {
                "rows": frame.height,
                "columns": frame.columns,
                "from": str(frame.select(pl.col("^(logged_at|measured_at|started_at)$")).min().row(0)[0]),
                "to": str(frame.select(pl.col("^(logged_at|measured_at|started_at)$")).max().row(0)[0]),
            }
            if any(c in frame.columns for c in ("logged_at", "measured_at", "started_at"))
            else {"rows": frame.height, "columns": frame.columns}
            for name, frame in tables.items()
        },
    }
    (args.out / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n")

    for name, frame in tables.items():
        log.info("%-20s %6d rows -> %s", name, frame.height, args.out / name / f"{name}.parquet")
    log.info(
        "implied height %sm; %d filler rows skipped; %d weigh-in rows routed to body_metrics",
        height_m,
        skipped,
        weigh_ins,
    )


if __name__ == "__main__":
    main()
